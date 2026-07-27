from __future__ import annotations

from datetime import date
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from clients.schneider.business import (
    load_match_data,
    merge_result_report,
    money,
    parse_consignment_range,
)


def _write_match_excel(path: Path, rows: list[list]):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["说明", None, None, None])
    sheet.append(["发票备注栏", "收货日期", "订单号", "行号"])
    for row in rows:
        sheet.append(row)
    workbook.save(path)
    workbook.close()


def test_non_consignment_excel_exact_match(tmp_path):
    path = tmp_path / "PP-AVXP—测试对账结果.xlsx"
    _write_match_excel(
        path,
        [
            ["非寄售批次A", "2026-07-20", "PO001", 10],
            ["非寄售批次A", "2026-07-22", "PO002", 20.0],
            ["非寄售批次B", "2026-07-23", "PO003", 30],
        ],
    )

    result = load_match_data(path, "非寄售批次A")

    assert result.company_code == "AVXP"
    assert result.invoice_kind == "non_consignment"
    assert result.receipt_start == date(2026, 7, 20)
    assert result.receipt_end == date(2026, 7, 22)
    assert result.row_keys == {("PO001", "10"), ("PO002", "20")}


def test_duplicate_order_line_is_rejected(tmp_path):
    path = tmp_path / "PP-AVXP—测试.xlsx"
    _write_match_excel(
        path,
        [
            ["非寄售批次A", "2026-07-20", "PO001", 10],
            ["非寄售批次A", "2026-07-21", "PO001", 10],
        ],
    )

    with pytest.raises(ValueError, match="重复"):
        load_match_data(path, "非寄售批次A")


def test_consignment_range():
    assert parse_consignment_range("寄售P厂2026.07.10-2026.07.11") == (
        date(2026, 7, 10),
        date(2026, 7, 11),
    )


def test_zero_amount_is_valid_money():
    assert money(0) == 0


def test_daily_report_merge_is_idempotent(tmp_path):
    source = tmp_path / "下载报告.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["发票号", "结果"])
    sheet.append(["INV001", "成功"])
    workbook.save(source)
    workbook.close()

    target, first_count = merge_result_report(source, tmp_path, date(2026, 7, 26))
    _, second_count = merge_result_report(source, tmp_path, date(2026, 7, 26))

    assert target.name == "施耐德客户系统发票录入结果=20260726.xlsx"
    assert first_count == 1
    assert second_count == 0
    merged = load_workbook(target, read_only=True)
    assert merged.active.max_row == 2
    merged.close()
