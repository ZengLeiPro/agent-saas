"""施耐德发票 POC 的业务解析与 Excel 处理。"""

from __future__ import annotations

import re
import shutil
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path


HEADER_ALIASES = {
    "remark": ("发票备注栏", "发票备注", "备注"),
    "receipt_date": ("收货日期", "收货日", "收货时间"),
    "order_no": ("订单号", "采购订单号", "po号", "po"),
    "line_no": ("行号", "订单行号", "po行号"),
}


def normalize_text(value) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", "", str(value)).strip()


def normalize_key(value) -> str:
    text = normalize_text(value)
    return re.sub(r"\.0$", "", text)


def money(value) -> Decimal:
    text = "" if value is None else str(value).replace(",", "").strip()
    try:
        return Decimal(text)
    except InvalidOperation as exc:
        raise ValueError(f"无法解析金额：{value!r}") from exc


def parse_invoice_kind(remark: str) -> str:
    compact = normalize_text(remark)
    if compact.startswith("非寄售"):
        return "non_consignment"
    if compact.startswith("寄售"):
        return "consignment"
    raise ValueError(f"发票备注未以“非寄售”或“寄售”开头：{remark!r}")


def _parse_date_text(value: str) -> date:
    compact = value.strip()
    for fmt in ("%Y.%m.%d", "%Y-%m-%d", "%Y/%m/%d", "%Y%m%d"):
        try:
            return datetime.strptime(compact, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"无法解析日期：{value!r}")


def parse_consignment_range(remark: str) -> tuple[date, date]:
    match = re.search(
        r"(\d{4}[./-]\d{1,2}[./-]\d{1,2})\s*(?:-|—|~|～|至)\s*"
        r"(\d{4}[./-]\d{1,2}[./-]\d{1,2})",
        remark,
    )
    if not match:
        raise ValueError(f"寄售备注中没有日期区间：{remark!r}")
    start = _parse_date_text(match.group(1))
    end = _parse_date_text(match.group(2))
    if start > end:
        raise ValueError(f"寄售日期区间倒置：{start} > {end}")
    return start, end


def extract_company_code(filename: str, reconciliation_name: str = "") -> str:
    for source in (Path(filename).stem, reconciliation_name):
        for candidate in re.findall(r"(?<![A-Z])[A-Z]{4}(?![A-Z])", source.upper()):
            if candidate not in {"XLSX", "HTTP", "HTTPS"}:
                return candidate
    raise ValueError(
        f"无法从 Excel 文件名或客户系统对账名称提取施耐德公司代码："
        f"{filename!r} / {reconciliation_name!r}"
    )


@dataclass(frozen=True)
class MatchRow:
    order_no: str
    line_no: str
    receipt_date: date
    values: tuple


@dataclass(frozen=True)
class MatchData:
    path: Path
    company_code: str
    invoice_kind: str
    rows: tuple[MatchRow, ...]
    receipt_start: date
    receipt_end: date

    @property
    def row_keys(self) -> set[tuple[str, str]]:
        return {(row.order_no, row.line_no) for row in self.rows}


def load_match_data(
    excel_path: str | Path,
    invoice_remark: str,
    reconciliation_name: str = "",
) -> MatchData:
    path = Path(excel_path)
    matrix, datemode = _read_match_matrix(path)
    header_index, columns = _find_header(matrix)
    expected_remark = normalize_text(invoice_remark)
    matched = []

    for values in matrix[header_index + 1 :]:
        if not any(value not in (None, "") for value in values):
            continue
        row_remark = normalize_text(_cell(values, columns["remark"]))
        if row_remark != expected_remark:
            continue
        receipt = _cell_date(_cell(values, columns["receipt_date"]), datemode)
        matched.append(
            MatchRow(
                order_no=normalize_key(_cell(values, columns["order_no"])),
                line_no=normalize_key(_cell(values, columns["line_no"])),
                receipt_date=receipt,
                values=tuple(values),
            )
        )

    if not matched:
        raise ValueError(
            f"对账 Excel 中没有“发票备注栏”严格匹配 {invoice_remark!r} 的数据。"
        )
    if any(not row.order_no or not row.line_no for row in matched):
        raise ValueError("对账 Excel 匹配行存在空订单号或空行号，禁止继续。")
    row_keys = [(row.order_no, row.line_no) for row in matched]
    if len(row_keys) != len(set(row_keys)):
        raise ValueError("对账 Excel 匹配行存在重复的“订单号+行号”，禁止猜测去重。")

    dates = [row.receipt_date for row in matched]
    return MatchData(
        path=path,
        company_code=extract_company_code(path.name, reconciliation_name),
        invoice_kind=parse_invoice_kind(invoice_remark),
        rows=tuple(matched),
        receipt_start=min(dates),
        receipt_end=max(dates),
    )


def _read_matrix(path: Path) -> tuple[list[list], int | None]:
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        rows = [list(row) for row in sheet.iter_rows(values_only=True)]
        workbook.close()
        return rows, None
    if suffix == ".xls":
        import xlrd

        workbook = xlrd.open_workbook(path)
        sheet = workbook.sheet_by_index(0)
        rows = [sheet.row_values(index) for index in range(sheet.nrows)]
        return rows, workbook.datemode
    raise ValueError(f"只支持 .xlsx/.xls 对账文件：{path}")


def _read_match_matrix(path: Path) -> tuple[list[list], int | None]:
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            for sheet in workbook.worksheets:
                rows = [list(row) for row in sheet.iter_rows(values_only=True)]
                try:
                    _find_header(rows)
                except ValueError:
                    continue
                return rows, None
        finally:
            workbook.close()
    elif suffix == ".xls":
        import xlrd

        workbook = xlrd.open_workbook(path)
        for sheet in workbook.sheets():
            rows = [sheet.row_values(index) for index in range(sheet.nrows)]
            try:
                _find_header(rows)
            except ValueError:
                continue
            return rows, workbook.datemode
    else:
        raise ValueError(f"只支持 .xlsx/.xls 对账文件：{path}")
    raise ValueError(
        "对账 Excel 所有工作表的前 30 行均未找到完整表头："
        "发票备注栏/收货日期/订单号/行号。"
    )


def _find_header(matrix: list[list]) -> tuple[int, dict[str, int]]:
    normalized_aliases = {
        key: tuple(normalize_text(alias).lower() for alias in aliases)
        for key, aliases in HEADER_ALIASES.items()
    }
    for row_index, values in enumerate(matrix[:30]):
        columns = {}
        ranks = {}
        for column_index, value in enumerate(values):
            header = normalize_text(value).lower()
            for key, aliases in normalized_aliases.items():
                if header in aliases:
                    rank = aliases.index(header)
                    if key not in ranks or rank < ranks[key]:
                        columns[key] = column_index
                        ranks[key] = rank
        if columns.keys() >= HEADER_ALIASES.keys():
            return row_index, columns
    raise ValueError(
        "对账 Excel 前 30 行未找到完整表头：发票备注栏/收货日期/订单号/行号。"
    )


def _cell(values: list | tuple, index: int):
    return values[index] if index < len(values) else None


def _cell_date(value, datemode: int | None) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)) and datemode is not None:
        import xlrd

        return xlrd.xldate_as_datetime(value, datemode).date()
    return _parse_date_text(str(value))


def merge_result_report(
    downloaded_report: str | Path,
    target_directory: str | Path,
    report_date: date,
) -> tuple[Path, int]:
    """把施耐德下载报表幂等合并到当日结果工作簿。"""
    source = Path(downloaded_report)
    target_dir = Path(target_directory)
    target_dir.mkdir(parents=True, exist_ok=True)
    target = target_dir / f"施耐德客户系统发票录入结果={report_date:%Y%m%d}.xlsx"

    rows, _ = _read_matrix(source)
    rows = [row for row in rows if any(value not in (None, "") for value in row)]
    if not rows:
        raise ValueError(f"下载报表为空：{source}")

    if not target.exists() and source.suffix.lower() == ".xlsx":
        shutil.copy2(source, target)
        return target, max(0, len(rows) - 1)

    from openpyxl import Workbook, load_workbook

    if target.exists():
        workbook = load_workbook(target)
        sheet = workbook.active
        target_header = [cell.value for cell in sheet[1]]
        if [normalize_text(x) for x in target_header] != [
            normalize_text(x) for x in rows[0]
        ]:
            workbook.close()
            raise ValueError("下载报表与当日结果工作簿表头不一致，禁止盲目合并。")
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "录入结果"
        sheet.append(rows[0])

    existing = {
        tuple(cell.value for cell in row)
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row)
    }
    appended = 0
    for row in rows[1:]:
        key = tuple(row)
        if key in existing:
            continue
        sheet.append(list(row))
        existing.add(key)
        appended += 1
    workbook.save(target)
    workbook.close()
    return target, appended
